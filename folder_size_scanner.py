"""
Folder Size Scanner
A GUI utility to find large files and folders and export them to Excel.
"""

import ctypes
import ctypes.wintypes
import os
import platform
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path

def _ensure_package(module_name, pip_name=None):
    try:
        return __import__(module_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name or module_name, "-q"])
        return __import__(module_name)


_ensure_package("openpyxl")
_ensure_package("customtkinter")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import customtkinter as ctk


# ─────────────────────────────────────────────
#  Scanning logic
# ─────────────────────────────────────────────

def get_folder_size(folder_path):
    """Recursively compute size of a folder in bytes."""
    total = 0
    try:
        for entry in os.scandir(folder_path):
            try:
                if entry.is_file(follow_symlinks=False):
                    total += get_disk_size(entry.path)
                elif entry.is_dir(follow_symlinks=False):
                    total += get_folder_size(entry.path)
            except (PermissionError, OSError):
                pass
    except (PermissionError, OSError):
        pass
    return total


def bytes_to_mb(b):
    return b / (1024 * 1024)


def bytes_to_human(b):
    if b >= 1024 ** 3:
        return f"{b / 1024**3:.2f} GB"
    if b >= 1024 ** 2:
        return f"{b / 1024**2:.2f} MB"
    if b >= 1024:
        return f"{b / 1024:.2f} KB"
    return f"{b} B"


def get_disk_size(file_path):
    """Return actual bytes allocated on disk, matching Windows Explorer's 'Size on Disk'.
    Handles cloud placeholders (OneDrive), NTFS-compressed, and sparse files.
    Falls back to logical size on non-Windows or on API error."""
    if platform.system() != "Windows":
        return os.path.getsize(file_path)
    high = ctypes.wintypes.DWORD(0)
    low = ctypes.windll.kernel32.GetCompressedFileSizeW(
        ctypes.c_wchar_p(file_path), ctypes.byref(high)
    )
    if low == 0xFFFFFFFF and ctypes.GetLastError() != 0:
        return os.path.getsize(file_path)
    return (high.value << 32) | low


def get_modified_date(path):
    try:
        ts = os.path.getmtime(path)
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
    except OSError:
        return "N/A"


def scan_directory(root_path, threshold_mb, progress_callback, stop_event):
    """
    Scan root_path and return (large_files, large_folders).
    progress_callback(message) is called during scanning.
    """
    threshold_bytes = threshold_mb * 1024 * 1024
    large_files = []
    large_folders = []

    progress_callback(f"Scanning: {root_path}")

    for dirpath, dirnames, filenames in os.walk(root_path, followlinks=False):
        if stop_event.is_set():
            break

        progress_callback(f"Checking: {dirpath}")

        # Check files
        for fname in filenames:
            if stop_event.is_set():
                break
            fpath = os.path.join(dirpath, fname)
            try:
                size = get_disk_size(fpath)
                if size >= threshold_bytes:
                    rel = os.path.relpath(fpath, root_path)
                    ext = Path(fpath).suffix.lower() or "(no ext)"
                    large_files.append({
                        "name": fname,
                        "relative_path": rel,
                        "full_path": fpath,
                        "size_bytes": size,
                        "size_mb": bytes_to_mb(size),
                        "size_human": bytes_to_human(size),
                        "extension": ext,
                        "modified": get_modified_date(fpath),
                        "parent_folder": dirpath,
                    })
            except (PermissionError, OSError):
                pass

        # Check sub-folders (not the root itself)
        for dname in dirnames:
            if stop_event.is_set():
                break
            dpath = os.path.join(dirpath, dname)
            try:
                size = get_folder_size(dpath)
                if size >= threshold_bytes:
                    rel = os.path.relpath(dpath, root_path)
                    large_folders.append({
                        "name": dname,
                        "relative_path": rel,
                        "full_path": dpath,
                        "size_bytes": size,
                        "size_mb": bytes_to_mb(size),
                        "size_human": bytes_to_human(size),
                        "modified": get_modified_date(dpath),
                        "parent_folder": dirpath,
                    })
            except (PermissionError, OSError):
                pass

    large_files.sort(key=lambda x: x["size_bytes"], reverse=True)
    large_folders.sort(key=lambda x: x["size_bytes"], reverse=True)
    return large_files, large_folders


# ─────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # Dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # Medium blue
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")   # Light blue tint
FILE_HDR_FILL = PatternFill("solid", fgColor="1F4E79")   # Deep blue (files)
FOLD_HDR_FILL = PatternFill("solid", fgColor="375623")   # Dark green (folders)
ALT_GREEN     = PatternFill("solid", fgColor="EBF5E1")   # Light green tint
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

THIN = Side(style="thin", color="BDD7EE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_font(size=11, color="FFFFFF"):
    return Font(name="Arial", bold=True, size=size, color=color)


def cell_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)


def apply_header_row(ws, row_num, values, fill, font_color="FFFFFF", heights=18):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = hdr_font(10, font_color)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row_num].height = heights


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_summary_sheet(ws, root_path, threshold_mb, large_files, large_folders, scan_time):
    ws.title = "📊 Summary"

    # Title banner
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Folder Size Scanner — Report"
    t.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   Scan time: {scan_time:.1f}s"
    sub.font = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    sub.fill = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Info rows
    info = [
        ("Scanned Folder", root_path),
        ("Size Threshold", f"{threshold_mb:,.1f} MB"),
        ("Large Files Found", f"{len(large_files):,}"),
        ("Large Folders Found", f"{len(large_folders):,}"),
        ("Total Items Found", f"{len(large_files) + len(large_folders):,}"),
    ]
    if large_files:
        total_file_bytes = sum(f["size_bytes"] for f in large_files)
        info.append(("Total Size (Files)", bytes_to_human(total_file_bytes)))
    if large_folders:
        biggest = large_folders[0]
        info.append(("Largest Folder", f"{biggest['name']}  ({biggest['size_human']})"))
    if large_files:
        biggest_f = large_files[0]
        info.append(("Largest File", f"{biggest_f['name']}  ({biggest_f['size_human']})"))

    label_fill = PatternFill("solid", fgColor="D6E4F0")
    value_fill = PatternFill("solid", fgColor="FFFFFF")

    for i, (label, value) in enumerate(info, start=4):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=10, color="1F3864")
        lc.fill = label_fill
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = BORDER

        vc = ws.cell(row=i, column=2, value=value)
        vc.font = Font(name="Arial", size=10, color="000000")
        vc.fill = value_fill
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = BORDER

        ws.merge_cells(f"B{i}:D{i}")
        ws.row_dimensions[i].height = 16

    set_col_widths(ws, [28, 55, 15, 15])


def build_files_sheet(ws, large_files, threshold_mb):
    ws.title = "📄 Large Files"

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Large Files  (≥ {threshold_mb:,.1f} MB)"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = FILE_HDR_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["#", "File Name", "Extension", "Size on Disk", "Size on Disk (MB)", "Modified Date", "Relative Path", "Parent Folder"]
    apply_header_row(ws, 2, headers, FILE_HDR_FILL)

    for i, f in enumerate(large_files, 1):
        row = i + 2
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        vals = [
            i,
            f["name"],
            f["extension"],
            f["size_human"],
            round(f["size_mb"], 3),
            f["modified"],
            f["relative_path"],
            f["parent_folder"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = cell_font()
            c.fill = fill
            c.border = BORDER
            c.alignment = Alignment(vertical="center", indent=1)
            if col == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 5:
                c.number_format = "#,##0.000"
                c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[row].height = 15

    # Freeze header
    ws.freeze_panes = "A3"

    # Auto-filter
    if large_files:
        ws.auto_filter.ref = f"A2:H{len(large_files)+2}"

    # Color scale on MB column (E)
    if large_files:
        last = len(large_files) + 2
        ws.conditional_formatting.add(
            f"E3:E{last}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    set_col_widths(ws, [5, 34, 10, 12, 12, 18, 55, 50])


def build_folders_sheet(ws, large_folders, threshold_mb):
    ws.title = "📁 Large Folders"

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"Large Folders  (≥ {threshold_mb:,.1f} MB)"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = FOLD_HDR_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["#", "Folder Name", "Size on Disk", "Size on Disk (MB)", "Modified Date", "Relative Path", "Parent Folder"]
    apply_header_row(ws, 2, headers, FOLD_HDR_FILL)

    for i, f in enumerate(large_folders, 1):
        row = i + 2
        fill = ALT_GREEN if i % 2 == 0 else WHITE_FILL
        vals = [
            i,
            f["name"],
            f["size_human"],
            round(f["size_mb"], 3),
            f["modified"],
            f["relative_path"],
            f["parent_folder"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = cell_font()
            c.fill = fill
            c.border = BORDER
            c.alignment = Alignment(vertical="center", indent=1)
            if col == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 4:
                c.number_format = "#,##0.000"
                c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[row].height = 15

    ws.freeze_panes = "A3"

    if large_folders:
        ws.auto_filter.ref = f"A2:G{len(large_folders)+2}"
        last = len(large_folders) + 2
        ws.conditional_formatting.add(
            f"D3:D{last}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ),
        )

    set_col_widths(ws, [5, 34, 12, 12, 18, 55, 50])


def export_to_excel(root_path, threshold_mb, large_files, large_folders, scan_time, out_path):
    wb = Workbook()
    ws_summary = wb.active
    build_summary_sheet(ws_summary, root_path, threshold_mb, large_files, large_folders, scan_time)

    ws_files = wb.create_sheet()
    build_files_sheet(ws_files, large_files, threshold_mb)

    ws_folders = wb.create_sheet()
    build_folders_sheet(ws_folders, large_folders, threshold_mb)

    # Set tab colors
    wb["📊 Summary"].sheet_properties.tabColor = "1F3864"
    wb["📄 Large Files"].sheet_properties.tabColor = "1F4E79"
    wb["📁 Large Folders"].sheet_properties.tabColor = "375623"

    wb.active = wb["📊 Summary"]
    wb.save(out_path)


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────

# Treeview palette per appearance mode (ttk doesn't auto-follow customtkinter themes)
TREE_PALETTE = {
    "Light": {
        "bg": "#FFFFFF", "fg": "#1F3864", "field_bg": "#FFFFFF",
        "heading_bg": "#1F538D", "heading_fg": "#FFFFFF",
        "selected_bg": "#3B8ED0", "selected_fg": "#FFFFFF",
        "odd": "#FFFFFF", "even": "#EBF3FB",
    },
    "Dark": {
        "bg": "#2B2B2B", "fg": "#DCE4EE", "field_bg": "#2B2B2B",
        "heading_bg": "#1F538D", "heading_fg": "#FFFFFF",
        "selected_bg": "#14375E", "selected_fg": "#FFFFFF",
        "odd": "#2B2B2B", "even": "#333333",
    },
}


class ScannerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.title("Folder Size Scanner")
        self.geometry("1060x720")
        self.minsize(880, 580)

        self._large_files = []
        self._large_folders = []
        self._scan_time = 0.0
        self._stop_event = threading.Event()
        self._scanning = False
        self._file_paths = {}     # tree iid -> full path (Large Files tab)
        self._folder_paths = {}   # tree iid -> full path (Large Folders tab)

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_header()
        self._build_controls()
        self._build_tabs()
        self._build_progress()
        self._build_statusbar()

        self._apply_tree_theme()

    # ── UI construction ──────────────────────

    def _build_header(self):
        header = ctk.CTkFrame(self, corner_radius=0, height=64,
                              fg_color=("#1F3864", "#111827"))
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header, text="🔍  Folder Size Scanner",
            font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
            text_color="white",
        ).grid(row=0, column=0, padx=20, pady=14, sticky="w")

        right = ctk.CTkFrame(header, fg_color="transparent")
        right.grid(row=0, column=2, padx=16, pady=10, sticky="e")

        ctk.CTkLabel(
            right, text="Theme:", font=ctk.CTkFont(size=12),
            text_color="white",
        ).pack(side="left", padx=(0, 8))

        self._theme_switch = ctk.CTkSegmentedButton(
            right, values=["Light", "Dark", "System"],
            command=self._on_theme_change,
        )
        self._theme_switch.set("System")
        self._theme_switch.pack(side="left")

    def _build_controls(self):
        ctrl = ctk.CTkFrame(self, corner_radius=0)
        ctrl.grid(row=1, column=0, sticky="ew")
        ctrl.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(ctrl, text="Folder:",
                     font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=0, column=0, padx=(16, 6), pady=12, sticky="w")

        self._folder_var = tk.StringVar()
        ctk.CTkEntry(
            ctrl, textvariable=self._folder_var,
            placeholder_text="Choose a folder to scan…",
            font=ctk.CTkFont(size=12),
        ).grid(row=0, column=1, padx=(0, 6), pady=12, sticky="ew")

        ctk.CTkButton(
            ctrl, text="Browse…", width=90, command=self._browse_folder,
        ).grid(row=0, column=2, padx=(0, 24), pady=12)

        ctk.CTkLabel(ctrl, text="Min size:",
                     font=ctk.CTkFont(size=12, weight="bold")).grid(
            row=0, column=3, padx=(0, 6), pady=12, sticky="w")

        self._threshold_var = tk.StringVar(value="100")
        ctk.CTkEntry(
            ctrl, textvariable=self._threshold_var, width=70,
            font=ctk.CTkFont(size=12),
        ).grid(row=0, column=4, padx=(0, 4), pady=12)

        ctk.CTkLabel(ctrl, text="MB",
                     font=ctk.CTkFont(size=12)).grid(
            row=0, column=5, padx=(0, 20), pady=12)

        btn_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        btn_frame.grid(row=0, column=6, padx=(0, 16), pady=12)

        self._scan_btn = ctk.CTkButton(
            btn_frame, text="▶  Scan", width=100,
            fg_color="#2E8B57", hover_color="#246B44",
            command=self._start_scan,
        )
        self._scan_btn.pack(side="left", padx=(0, 6))

        self._stop_btn = ctk.CTkButton(
            btn_frame, text="⏹  Stop", width=90,
            fg_color="#B55419", hover_color="#8B3F11",
            state="disabled", command=self._stop_scan,
        )
        self._stop_btn.pack(side="left", padx=(0, 6))

        self._export_btn = ctk.CTkButton(
            btn_frame, text="📥  Export to Excel", width=170,
            state="disabled", command=self._export_excel,
        )
        self._export_btn.pack(side="left")

    def _build_tabs(self):
        self._tabs = ctk.CTkTabview(self, corner_radius=8)
        self._tabs.grid(row=2, column=0, sticky="nsew", padx=12, pady=(8, 4))

        files_tab = self._tabs.add("📄  Large Files")
        folders_tab = self._tabs.add("📁  Large Folders")
        log_tab = self._tabs.add("📋  Scan Log")

        self._files_tree = self._build_tree(
            files_tab,
            ["#", "File Name", "Extension", "Size on Disk",
             "Size on Disk (MB)", "Modified", "Relative Path"],
            [40, 260, 80, 110, 130, 140, 420],
        )
        self._folders_tree = self._build_tree(
            folders_tab,
            ["#", "Folder Name", "Size on Disk",
             "Size on Disk (MB)", "Modified", "Relative Path"],
            [40, 300, 110, 130, 140, 420],
        )
        self._log_text = self._build_log(log_tab)

    def _build_tree(self, parent, columns, col_widths):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=2, pady=2)
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        tree = ttk.Treeview(
            container, columns=columns, show="headings",
            style="Custom.Treeview",
        )
        for col, w in zip(columns, col_widths):
            tree.heading(col, text=col,
                         command=lambda c=col, t=tree: self._sort_tree(t, c, False))
            anchor = "e" if col in ("Size on Disk (MB)", "#") else "w"
            tree.column(col, width=w, anchor=anchor, minwidth=30)

        vsb = ctk.CTkScrollbar(container, command=tree.yview)
        hsb = ctk.CTkScrollbar(container, orientation="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree.bind("<Button-3>", lambda e, t=tree: self._show_context_menu(e, t))
        tree.bind("<Double-Button-1>", lambda e, t=tree: self._reveal_selected(t))

        return tree

    def _build_log(self, parent):
        txt = ctk.CTkTextbox(
            parent, font=ctk.CTkFont(family="Consolas", size=11),
            wrap="none",
        )
        txt.pack(fill="both", expand=True, padx=2, pady=2)
        txt.configure(state="disabled")
        return txt

    def _build_progress(self):
        self._progress = ctk.CTkProgressBar(self, mode="indeterminate", height=6)
        # Hidden initially; shown during scans via grid() / grid_remove()

    def _build_statusbar(self):
        self._status_var = tk.StringVar(value="Ready — choose a folder and click Scan.")
        bar = ctk.CTkFrame(self, corner_radius=0, height=28,
                           fg_color=("#E3E8EF", "#1C1F23"))
        bar.grid(row=4, column=0, sticky="ew")
        ctk.CTkLabel(
            bar, textvariable=self._status_var,
            anchor="w", font=ctk.CTkFont(size=11),
        ).pack(fill="x", padx=16, pady=4)

    # ── Theme ────────────────────────────────

    def _on_theme_change(self, choice):
        ctk.set_appearance_mode(choice.lower())
        self._apply_tree_theme()

    def _apply_tree_theme(self):
        mode = ctk.get_appearance_mode()  # resolves "System" to "Light" or "Dark"
        p = TREE_PALETTE.get(mode, TREE_PALETTE["Light"])

        style = ttk.Style(self)
        try:
            style.theme_use("default")
        except tk.TclError:
            pass
        style.configure(
            "Custom.Treeview",
            font=("Segoe UI", 10), rowheight=24,
            background=p["bg"], fieldbackground=p["field_bg"],
            foreground=p["fg"], borderwidth=0,
        )
        style.configure(
            "Custom.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background=p["heading_bg"], foreground=p["heading_fg"],
            relief="flat",
        )
        style.map(
            "Custom.Treeview",
            background=[("selected", p["selected_bg"])],
            foreground=[("selected", p["selected_fg"])],
        )
        style.map(
            "Custom.Treeview.Heading",
            background=[("active", p["heading_bg"])],
        )

        for tree in (getattr(self, "_files_tree", None),
                     getattr(self, "_folders_tree", None)):
            if tree is not None:
                tree.tag_configure("odd", background=p["odd"], foreground=p["fg"])
                tree.tag_configure("even", background=p["even"], foreground=p["fg"])

    # ── Sorting ──────────────────────────────

    def _sort_tree(self, tree, col, reverse):
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        try:
            data.sort(key=lambda t: float(t[0].replace(",", "")), reverse=reverse)
        except ValueError:
            data.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (_, k) in enumerate(data):
            tree.move(k, "", index)
            tag = "even" if index % 2 == 0 else "odd"
            tree.item(k, tags=(tag,))
        tree.heading(col, command=lambda: self._sort_tree(tree, col, not reverse))

    # ── Actions ──────────────────────────────

    def _browse_folder(self):
        path = filedialog.askdirectory(title="Choose a folder to scan")
        if path:
            self._folder_var.set(path)

    # ── Reveal in file manager ───────────────

    def _show_context_menu(self, event, tree):
        iid = tree.identify_row(event.y)
        if not iid:
            return
        tree.selection_set(iid)
        tree.focus(iid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Show in Explorer",
                         command=lambda: self._reveal_selected(tree))
        menu.add_command(label="Open Containing Folder",
                         command=lambda: self._open_parent(tree))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _selected_path(self, tree):
        sel = tree.selection()
        if not sel:
            return None, None
        is_folder = tree is self._folders_tree
        paths = self._folder_paths if is_folder else self._file_paths
        return paths.get(sel[0]), is_folder

    def _reveal_selected(self, tree):
        path, is_folder = self._selected_path(tree)
        if not path:
            return
        if not os.path.exists(path):
            messagebox.showwarning(
                "Path Not Found",
                f"This path no longer exists:\n{path}",
            )
            return
        self._reveal_path(path, is_folder)

    def _open_parent(self, tree):
        path, _ = self._selected_path(tree)
        if not path:
            return
        parent = os.path.dirname(path)
        if not os.path.isdir(parent):
            messagebox.showwarning(
                "Folder Not Found",
                f"Parent folder no longer exists:\n{parent}",
            )
            return
        self._reveal_path(parent, is_folder=True)

    def _reveal_path(self, path, is_folder):
        """Show path in the OS file manager.

        Files are revealed with the item selected in the parent folder.
        Folders are opened so their contents are visible."""
        system = platform.system()
        try:
            if system == "Windows":
                if is_folder:
                    os.startfile(path)
                else:
                    subprocess.Popen(
                        f'explorer /select,"{os.path.normpath(path)}"'
                    )
            elif system == "Darwin":
                if is_folder:
                    subprocess.Popen(["open", path])
                else:
                    subprocess.Popen(["open", "-R", path])
            else:
                target = path if is_folder else os.path.dirname(path)
                subprocess.Popen(["xdg-open", target])
        except Exception as e:
            messagebox.showerror("Reveal Error", f"Could not open:\n{e}")

    def _start_scan(self):
        folder = self._folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Invalid Folder", "Please choose a valid folder first.")
            return
        try:
            threshold = float(self._threshold_var.get())
            if threshold <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid Threshold",
                                 "Please enter a positive number for the size threshold.")
            return

        for tree in (self._files_tree, self._folders_tree):
            for item in tree.get_children():
                tree.delete(item)
        self._file_paths.clear()
        self._folder_paths.clear()
        self._log_clear()
        self._large_files = []
        self._large_folders = []
        self._export_btn.configure(state="disabled")
        self._stop_event.clear()
        self._scanning = True

        self._scan_btn.configure(state="disabled")
        self._stop_btn.configure(state="normal")
        self._progress.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 4))
        self._progress.start()
        self._status("Scanning…")

        self._scan_start_time = datetime.now()
        threading.Thread(
            target=self._scan_worker, args=(folder, threshold), daemon=True,
        ).start()

    def _scan_worker(self, folder, threshold):
        try:
            files, folders = scan_directory(
                folder, threshold,
                progress_callback=lambda msg: self.after(0, self._log_append, msg),
                stop_event=self._stop_event,
            )
            elapsed = (datetime.now() - self._scan_start_time).total_seconds()
            self.after(0, self._scan_done, files, folders, folder, threshold, elapsed)
        except Exception as e:
            self.after(0, self._scan_error, str(e))

    def _scan_done(self, files, folders, folder, threshold, elapsed):
        self._large_files = files
        self._large_folders = folders
        self._scan_time = elapsed
        self._scanning = False
        self._progress.stop()
        self._progress.grid_remove()
        self._scan_btn.configure(state="normal")
        self._stop_btn.configure(state="disabled")

        for i, f in enumerate(files):
            tag = "even" if i % 2 == 0 else "odd"
            iid = self._files_tree.insert("", "end", tags=(tag,), values=(
                i + 1, f["name"], f["extension"],
                f["size_human"], f"{f['size_mb']:.3f}",
                f["modified"], f["relative_path"],
            ))
            self._file_paths[iid] = f["full_path"]

        for i, f in enumerate(folders):
            tag = "even" if i % 2 == 0 else "odd"
            iid = self._folders_tree.insert("", "end", tags=(tag,), values=(
                i + 1, f["name"],
                f["size_human"], f"{f['size_mb']:.3f}",
                f["modified"], f["relative_path"],
            ))
            self._folder_paths[iid] = f["full_path"]

        stopped = " (scan stopped early)" if self._stop_event.is_set() else ""
        self._status(
            f"✔  Done{stopped} — {len(files)} large file(s), {len(folders)} large folder(s) "
            f"found above {threshold:.1f} MB  |  Scan time: {elapsed:.1f}s"
        )
        if files or folders:
            self._export_btn.configure(state="normal")
        self._log_append(f"\n✔ Scan complete. {len(files)} files, {len(folders)} folders.")

    def _scan_error(self, msg):
        self._scanning = False
        self._progress.stop()
        self._progress.grid_remove()
        self._scan_btn.configure(state="normal")
        self._stop_btn.configure(state="disabled")
        messagebox.showerror("Scan Error", f"An error occurred:\n{msg}")
        self._status("Error during scan.")

    def _stop_scan(self):
        self._stop_event.set()
        self._status("Stopping scan…")
        self._stop_btn.configure(state="disabled")

    def _export_excel(self):
        folder = self._folder_var.get().strip()
        threshold = float(self._threshold_var.get())
        default_name = f"FolderScan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = filedialog.asksaveasfilename(
            title="Save Excel Report",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not out_path:
            return
        try:
            export_to_excel(folder, threshold, self._large_files, self._large_folders,
                            self._scan_time, out_path)
            self._status(f"✔  Excel report saved: {out_path}")
            if messagebox.askyesno("Export Complete",
                                   f"Report saved to:\n{out_path}\n\nOpen the file now?"):
                if platform.system() == "Windows":
                    os.startfile(out_path)
                elif platform.system() == "Darwin":
                    subprocess.Popen(["open", out_path])
                else:
                    subprocess.Popen(["xdg-open", out_path])
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel:\n{e}")

    # ── Helpers ──────────────────────────────

    def _status(self, msg):
        self._status_var.set(msg)

    def _log_append(self, msg):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", msg + "\n")
        self._log_text.see("end")
        self._log_text.configure(state="disabled")

    def _log_clear(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    app = ScannerApp()
    app.mainloop()
